import os
import uuid
import csv
from typing import List, Optional, Tuple
from pathlib import Path

from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook

from langchain_text_splitters import RecursiveCharacterTextSplitter
from app.core.config import settings


def get_file_extension(filename: str) -> str:
    return Path(filename).suffix.lower().replace(".", "")


def extract_text_from_pdf(file_path: str) -> Tuple[str, int]:
    reader = PdfReader(file_path)
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text_parts.append(page_text)
    return "\n\n".join(text_parts), len(reader.pages)


def extract_text_from_docx(file_path: str) -> Tuple[str, int]:
    doc = DocxDocument(file_path)
    text_parts = [para.text for para in doc.paragraphs if para.text.strip()]
    full_text = "\n".join(text_parts)
    page_count = max(1, len(full_text) // 3000)
    return full_text, page_count


def extract_text_from_txt(file_path: str) -> Tuple[str, int]:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    page_count = max(1, len(text) // 3000)
    return text, page_count


def extract_text_from_csv(file_path: str) -> Tuple[str, int]:
    text_parts = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            text_parts.append(" | ".join(row))
    return "\n".join(text_parts), 1


def extract_text_from_excel(file_path: str) -> Tuple[str, int]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    text_parts = []
    for sheet_name in wb.sheetnames:
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
            if row_text.strip():
                text_parts.append(row_text)
    wb.close()
    return "\n".join(text_parts), len(wb.sheetnames)


def extract_text(file_path: str, file_type: str) -> Tuple[str, int]:
    extractors = {
        "pdf": extract_text_from_pdf,
        "docx": extract_text_from_docx,
        "doc": extract_text_from_docx,
        "txt": extract_text_from_txt,
        "md": extract_text_from_txt,
        "markdown": extract_text_from_txt,
        "csv": extract_text_from_csv,
        "xlsx": extract_text_from_excel,
        "xls": extract_text_from_excel,
    }
    extractor = extractors.get(file_type)
    if not extractor:
        raise ValueError(f"Unsupported file type: {file_type}")
    return extractor(file_path)


def clean_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    text = "\n".join(lines)
    while "  " in text:
        text = text.replace("  ", " ")
    return text.strip()


def split_text(text: str, chunk_size: int = None, chunk_overlap: int = None) -> List[str]:
    chunk_size = chunk_size or settings.CHUNK_SIZE
    chunk_overlap = chunk_overlap or settings.CHUNK_OVERLAP
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        length_function=len,
        separators=["\n\n", "\n", ". ", " ", ""]
    )
    return splitter.split_text(text)


def generate_unique_filename(original_filename: str) -> str:
    ext = Path(original_filename).suffix
    unique_id = uuid.uuid4().hex[:12]
    return f"{unique_id}{ext}"


def save_uploaded_file(file_content: bytes, original_filename: str) -> Tuple[str, str]:
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    saved_filename = generate_unique_filename(original_filename)
    full_path = os.path.join(settings.UPLOAD_DIR, saved_filename)
    with open(full_path, "wb") as f:
        f.write(file_content)
    return saved_filename, full_path
